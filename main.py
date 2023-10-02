import requests
import openpyxl
import json

# Открываем файл и загружаем данные из JSON
with open('credentials.json', 'r') as json_file:
    data = json.load(json_file)

# Извлекаем значения ключей и сохраняем их в переменные
msisdn = data.get('msisdn')
template_id_list_from_json = data.get('template_id')
# token = data.get('token')

# Выводим значения переменных
print("msisdn:", msisdn)
print("template_id:", template_id_list_from_json)
# print("token:", token)
print("################################################################################")
# Объявляем список с темплейтами и номер, на который отправляется

template_id_var = template_id_list_from_json
recipient_msisdn_var = msisdn

# создаем классы и функции для отправки СМС
class MessageSender:
    def __init__(self, url_post, headers_post):
        self.url_post = url_post
        self.headers_post = headers_post

    def send_message(self, data_post):
        response = requests.post(self.url_post, headers=self.headers_post, json=data_post)
        return response

class ResponseHandler:
    def __init__(self, response):
        self.response = response

    def get_status(self):
        response_data = self.response.json().get('data', {})
        status = response_data.get('attributes', {}).get('status')
        return status

# урл и хедеры для отправки (в данном случае используем моковские)
url_post = "https://2eab44fc-e0a2-4c43-9c55-d4bbc627e72c.mock.pstmn.io/api/v1/messages/"
headers_post = {
    "Content-Type": "application/json",
    "User-Agent": "PostmanRuntime/7.32.3",
    "Accept": "*/*",
    "Accept-Encoding": "gzip, deflate, br"
}
# Создаем Класс
message_sender = MessageSender(url_post, headers_post)

# создаем бади для пост запроса, где меняется темплейт СМС
for template_id in template_id_var:
    data_post = {
        "data": {
            "type": "messages",
            "attributes": {
                "request-id": "69099001-7dc3-11e7-80c3-005565f",
                "originator": "oleksii.kariaka",
                "language": "ukr",
                "sms-recipient": recipient_msisdn_var,
                "template-id": template_id,
                "params": {
                    "recipient_msisdn": recipient_msisdn_var
                }
            }
        }
    }
    # отправляем запрос с необходимым темплейтом
    response = message_sender.send_message(data_post)
    response_handler = ResponseHandler(response)
    status = response_handler.get_status()


    # проверка и вывод сообщени об отправке в консоль
    if status is not None:
        print("Статус код запроса POST SMS 202 - успешно")
        print(f"СМС с темплейтом {template_id} было отправленно")
        print("Статус:", status)
        print("################################################################################")
    else:
        print("Статус не найден в ответе.")


# вызываем метод для отправки
response = message_sender.send_message(data_post)

"""Получаем тексты СМС"""
# задаем урлу и хедеры (используем мок)
url_get = "https://2eab44fc-e0a2-4c43-9c55-d4bbc627e72c.mock.pstmn.io/api/v1/messages?filter%5Brecipient-msisdn%5D=380981402174"
headers_get = {
    "User-Agent": "PostmanRuntime/7.32.3",
    "Accept": "*/*",
    "Accept-Encoding": "gzip, deflate, br"
}
# получаем респонс из мока
response_get = requests.get(url_get, headers=headers_get)
response_get_json = response_get.json()

# проверяем статус код

if response_get.status_code == 200:
    print("Запрос GET SMS успешен - 200")
    # print(response_get.json())  # Вывести ответ в формате JSON
else:
   print("Ошибка запроса")

# создаем список, в котором будут словари "Template_id": "content"
content_sms_list = {}
# получаем из json тексты СМС
for item in response_get_json.get("data", []):
    attributes = item.get("attributes", {})
    payloads = attributes.get("payloads", [])
    for payload in payloads:
        template_id = payload.get("template-id")
        if template_id in template_id_var:
            content = payload.get("content")
            # print(f"Template ID: {template_id}, Content: {content}")
            content_sms_list[template_id] = content

print(content_sms_list)

"""Создаем список из файла Эксель для того, чтобы сравнить его с полученным (content_sms_list)"""

# Загрузка данных из файла Excel
file_path = "C:\\Users\\Karyaka\\PycharmProjects\\automationLesson\\Template_Test.xlsx"  # укажите путь к вашему файлу Excel
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Создание пустого словаря для хранения результатов
template_content_dict = {}

# Получение заголовков из первой строки
header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

# Находим индексы столбцов с нужными названиями
template_id_col_idx = header_row.index("Template Id")
content_col_idx = header_row.index("sms-content-ukr")

# Пробегаем по всем строкам
for row in sheet.iter_rows(min_row=2, values_only=True):
    template_id = row[template_id_col_idx]
    content = row[content_col_idx]

    if template_id in template_id_var:
        template_content_dict[template_id] = content

# Вывод результата
# for template_id, content in template_content_dict.items():
#     print(f"Template Id: {template_id}, Content: {content}")

# print(template_content_dict)
print('################################################################################')



"""Сравниваем тексты из API и из файла Excell"""

# Получение пересечения ключей
common_template_ids = set(content_sms_list.keys()) & set(template_content_dict.keys())

# Вывод соответствующих значений для каждого общего ключа
for template_id in common_template_ids:

    content_sms = content_sms_list.get(template_id)
    content_template = template_content_dict.get(template_id)

    print(f"Template ID: {template_id}")
    print(f"Content from API: {content_sms}")
    print(f"Content from Excel: {content_template}")
    if content_sms == content_template:
        print('Тексты совпадают!!!!')
    else:
        print('Тексты НЕ совпадают :(')
    print("################################################################################")

    """Создаем список для сохранения результатов"""
    comparison_results = []

    """Сравниваем тексты из API и из файла Excell"""
    common_template_ids = set(content_sms_list.keys()) & set(template_content_dict.keys())

    for template_id in common_template_ids:
        content_sms = content_sms_list.get(template_id, "Не найден в API")
        content_template = template_content_dict.get(template_id, "Не найден в Excel")

        result = {
            "Template ID": template_id,
            "Текст из API": content_sms,
            "Текст из Excel": content_template
        }
        comparison_results.append(result)


    """Выгружаем данные в Excel файл"""
    result_file_path = "Comparison_Results.xlsx"
    result_workbook = openpyxl.Workbook()
    result_sheet = result_workbook.active
    result_sheet.title = "Comparison Results"

    result_sheet.append(["Template ID", "Текст из API", "Текст из Excel"])

    for result in comparison_results:
        result_sheet.append([result["Template ID"], result["Текст из API"], result["Текст из Excel"]])

    result_file_path = "Comparison_Results.xlsx"
    result_workbook.save(result_file_path)
print(f"Результаты сравнения сохранены в файл: {result_file_path}")
print('################################################################################')